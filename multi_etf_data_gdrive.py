"""
╔══════════════════════════════════════════════════════════════════════════════╗
║         Multi-ETF iNAV vs LTP Alert System  —  GitHub Actions Edition       ║
╚══════════════════════════════════════════════════════════════════════════════╝

WHAT THIS SCRIPT DOES:
──────────────────────
1. Downloads an Excel file from Google Drive that contains a list of ETFs.
2. For each ETF, fetches live market data:
      - LTP          : Last Traded Price  (from Yahoo Finance)
      - NAV / iNAV   : Net Asset Value    (from NSE / BSE / AMFI / mfapi)
      - Expense Ratio: Annual fund cost % (from mfdata.in / captnemo)
      - 52W High/Low : 52-week price range
      - Beta, P/E    : Risk and valuation metrics
      - Volume       : Number of units traded today
3. Calculates the % difference between LTP and NAV.
4. Sends an Email + SMS alert if the difference is within your threshold.
5. Writes all results back into the Excel file (3 sheets: data, PE history, PE comparison).
6. Uploads the updated Excel back to Google Drive.

HOW IT IS TRIGGERED:
────────────────────
This script is designed to be run by GitHub Actions on a schedule (every 2 hours).
It runs ONE cycle and exits — GitHub Actions handles the repeat scheduling.
You do NOT need your PC on. Everything runs in the cloud for free.

GITHUB SECRETS REQUIRED  (set these in your repo → Settings → Secrets):
──────────────────────────────────────────────────────────────────────────
  GDRIVE_FILE_ID               – The ID of your Excel file in Google Drive
  GDRIVE_SERVICE_ACCOUNT_JSON  – The full JSON key for your Google service account
  MY_EMAIL                     – Your Gmail address (used to send AND receive alerts)
  MY_EMAIL_PSWRD               – Your Gmail App Password (NOT your Gmail login password)
  TWILIO_ACCOUNT_SID_NEW       – Twilio account SID (for SMS alerts)
  TWILIO_NUMBER                – Twilio phone number to send SMS from
  TWILIO_TO_NUMBER_NEW         – Your phone number to receive SMS alerts
  TWILIO_AUTH_TOKEN_NEW        – Twilio auth token

EXCEL FILE LAYOUT  (Sheet named "ETF Data"):
────────────────────────────────────────────
  Column A – ETF Symbol    e.g. MAFANG        <- YOU fill this
  Column B – ISIN          e.g. INF769K01HF4  <- YOU fill this
  Column C – Name          e.g. Mirae Asset   <- optional, script fills if blank
  Column D – Alert Threshold %  e.g. 15.0     <- optional, defaults to 15.0
  Columns E onwards -> filled automatically by this script every run

DEPENDENCIES  (install with pip):
──────────────────────────────────
  pip install requests yfinance openpyxl google-api-python-client google-auth
"""

# ─────────────────────────────────────────────────────────────────────────────
# STANDARD LIBRARY IMPORTS
# These come built into Python — no installation needed.
# ─────────────────────────────────────────────────────────────────────────────

import json        # For reading/writing JSON data (used for Google credentials and cache)
import time        # For timestamps and sleep/wait functions
import smtplib     # For sending emails via Gmail's SMTP server
import logging     # For printing structured log messages with timestamps
import os          # For reading environment variables (secrets set in GitHub)
import io          # For handling file data streams in memory (used in Drive download)
from datetime import datetime, time as dtime   # datetime = date+time object, dtime = time-only
from pathlib import Path                        # A cleaner way to work with file paths

# ─────────────────────────────────────────────────────────────────────────────
# THIRD-PARTY LIBRARY IMPORTS
# These must be installed via pip (see dependencies above).
# ─────────────────────────────────────────────────────────────────────────────

import requests    # For making HTTP requests to APIs (NSE, BSE, AMFI, mfapi, etc.)
import yfinance as yf  # Yahoo Finance wrapper — gives us LTP, 52W high/low, P/E, Beta, Volume

import openpyxl    # For reading and writing .xlsx Excel files
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # Excel cell styling
from openpyxl.utils import get_column_letter  # Converts column number to letter (e.g. 5 -> "E")

# Google Drive API libraries
from googleapiclient.discovery import build             # Builds the Drive API client object
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload  # For upload/download
from google.oauth2 import service_account              # For authenticating with a service account


# ─────────────────────────────────────────────────────────────────────────────
# LOGGING SETUP
# ─────────────────────────────────────────────────────────────────────────────
# logging prints messages to the console with a timestamp and severity level.
# Levels in order: DEBUG -> INFO -> WARNING -> ERROR -> CRITICAL
# We set level=INFO, so DEBUG messages are hidden but everything else shows.
# In GitHub Actions, these logs appear in the workflow run output.
# ─────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",  # e.g. "2026-04-09 10:00:00  INFO  ..."
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)  # Creates a logger named after this file


# ─────────────────────────────────────────────────────────────────────────────
# ENVIRONMENT VARIABLES (SECRETS)
# ─────────────────────────────────────────────────────────────────────────────
# os.getenv("NAME") reads the value of an environment variable named NAME.
# In GitHub Actions, secrets are injected as environment variables via the
# workflow YAML's  env:  block.  .strip() removes accidental spaces/newlines.
# If a secret is missing, getenv returns "" (empty string) rather than crashing.
# ─────────────────────────────────────────────────────────────────────────────

DRIVE_FILE_ID = os.getenv("GDRIVE_FILE_ID", "").strip()
# ^ The Google Drive file ID of your Excel file.
# Found in the Drive share URL: drive.google.com/file/d/FILE_ID_HERE/view

GDRIVE_CREDS = os.getenv("GDRIVE_SERVICE_ACCOUNT_JSON", "").strip()
# ^ The full contents of the Google service account .json key file.
# This is how the script authenticates with Google Drive without a browser login.

my_email           = os.getenv("MY_EMAIL", "").strip()
my_password        = os.getenv("MY_EMAIL_PSWRD", "").strip()
# ^ Gmail address and App Password for sending alert emails.
# Use a Gmail App Password (not your normal Gmail password):
# Google Account -> Security -> 2-Step Verification -> App Passwords

twilio_account_sid = os.getenv("TWILIO_ACCOUNT_SID_NEW", "").strip()
twilio_number      = os.getenv("TWILIO_NUMBER", "").strip()
twilio_to_number   = os.getenv("TWILIO_TO_NUMBER_NEW", "").strip()
twilio_auth_token  = os.getenv("TWILIO_AUTH_TOKEN_NEW", "").strip()
# ^ Twilio credentials for sending SMS alerts.
# Sign up free at twilio.com — you get a trial number to send SMS from.


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION DICTIONARY
# ─────────────────────────────────────────────────────────────────────────────
# A single dict that holds all tunable settings in one place.
# If you want to change a setting, change it here — not scattered through the code.
# ─────────────────────────────────────────────────────────────────────────────

CONFIG = {
    # Local filename used while the script runs on the GitHub Actions runner.
    # The file is downloaded from Drive, processed, then uploaded back.
    "excel_file": "etf_data_amfi1.xlsx",

    # Local file to cache the last-fetched data per ETF symbol.
    # Useful for debugging — you can inspect what was last fetched.
    "cache_file": "etf_cache.json",

    # If column D (Alert Threshold) is blank for an ETF, use this default value.
    # An alert fires when abs(LTP - NAV) / NAV * 100 <= threshold.
    "default_threshold_pct": 15.0,

    # Email alert settings.
    # "enabled" auto-sets to True only if both email secrets are present.
    "email": {
        "enabled":         bool(my_email and my_password),
        "smtp_host":       "smtp.gmail.com",  # Gmail's outgoing mail server
        "smtp_port":       587,               # Port 587 = STARTTLS (encrypted)
        "sender_email":    my_email,
        "sender_password": my_password,
        "recipient_email": my_email,          # Sending to yourself
    },

    # SMS alert settings via Twilio.
    # "enabled" auto-sets to True only if Twilio secrets are present.
    "sms": {
        "enabled":     bool(twilio_account_sid and twilio_auth_token),
        "account_sid": twilio_account_sid,
        "auth_token":  twilio_auth_token,
        "from_number": twilio_number,
        "to_number":   twilio_to_number,
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL SHEET NAMES
# ─────────────────────────────────────────────────────────────────────────────
# Constants for the three sheets we write to in the Excel file.
# Using constants means if you rename a sheet, you only change it in one place.
# ─────────────────────────────────────────────────────────────────────────────

MAIN_SHEET    = "ETF Data"       # Main sheet: one row per ETF with all live data
PE_HIST_SHEET = "PE History"     # Accumulates P/E ratio snapshots over time
PE_CMP_SHEET  = "PE Comparison"  # Compares current P/E vs ~30 days ago


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT COLUMN HEADERS
# ─────────────────────────────────────────────────────────────────────────────
# These are the column headers written to the MAIN_SHEET starting from column E.
# The order here must match the order values are written in write_results_to_excel().
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_COLS = [
    "LTP (Rs)",               # Last Traded Price on NSE
    "Expense Ratio (%)",      # Annual Total Expense Ratio of the ETF
    "52W High (Rs)",          # Highest price in the last 52 weeks
    "52W Low (Rs)",           # Lowest price in the last 52 weeks
    "Rise from 52W Low (%)",  # How much % LTP has risen from the 52W low
    "Fall from 52W High (%)", # How much % LTP has fallen from the 52W high
    "Beta",                   # Volatility vs market (>1 = more volatile, <1 = less)
    "PEG Ratio",              # Price/Earnings-to-Growth (N/A for Indian ETFs)
    "P/E Ratio",              # Price-to-Earnings ratio
    "Volume",                 # Number of units traded today
    "NAV / iNAV (Rs)",        # Net Asset Value (declared) or iNAV (live, during market hours)
    "NAV Date",               # Date/source of the NAV used
    "Diff vs NAV (%)",        # (LTP - NAV) / NAV * 100 — positive=premium, negative=discount
    "Last Updated",           # Timestamp of this script run
]

# Column positions in the Excel sheet (1-indexed, A=1, B=2, etc.)
COL_SYMBOL       = 1   # A — ETF symbol (e.g. MAFANG)
COL_ISIN         = 2   # B — ISIN code
COL_NAME         = 3   # C — Fund name
COL_THRESHOLD    = 4   # D — Alert threshold %
COL_OUTPUT_START = 5   # E — First output column (LTP); subsequent columns follow in order


# =============================================================================
# SECTION 1: GOOGLE DRIVE INTEGRATION
# =============================================================================
# These functions handle downloading the Excel from Drive before the run,
# and uploading it back after the run completes.
#
# Authentication uses a "Service Account" — a special Google identity created
# in Google Cloud Console that can be used by scripts without a browser login.
# The service account's credentials are stored as a JSON file (GDRIVE_CREDS).
# =============================================================================

def _get_drive_service():
    """
    Authenticate with Google Drive and return a Drive API client object.

    HOW IT WORKS:
    - Reads the service account JSON from the environment variable GDRIVE_CREDS.
    - Parses the JSON and creates a credentials object with Drive access scope.
    - Builds and returns a 'service' object we use to call Drive API methods.

    RAISES RuntimeError with clear instructions if:
    - The secret is empty (not set in GitHub)
    - The secret is not valid JSON (pasted incorrectly)
    - The file ID secret is missing
    """

    # Check that the credentials secret is not empty
    if not GDRIVE_CREDS:
        raise RuntimeError(
            "GDRIVE_SERVICE_ACCOUNT_JSON secret is empty or not set.\n"
            "Fix: GitHub repo -> Settings -> Secrets and variables -> Actions\n"
            "     -> confirm GDRIVE_SERVICE_ACCOUNT_JSON exists and is not empty.\n"
            "     Also confirm the workflow env: block passes it to the run step."
        )

    # Try to parse the JSON string into a Python dictionary
    try:
        creds_dict = json.loads(GDRIVE_CREDS)
    except json.JSONDecodeError as e:
        # Show first 120 characters to help diagnose pasting issues (safe — no private key there)
        preview = repr(GDRIVE_CREDS[:120])
        raise RuntimeError(
            f"GDRIVE_SERVICE_ACCOUNT_JSON is not valid JSON: {e}\n"
            f"First 120 chars received: {preview}\n"
            "Fix: paste the ENTIRE contents of the service account .json file\n"
            "     as the secret value — no extra quotes, no truncation."
        ) from e

    # Check that the Drive file ID secret is also not empty
    if not DRIVE_FILE_ID:
        raise RuntimeError(
            "GDRIVE_FILE_ID secret is empty or not set.\n"
            "Fix: copy the file ID from the Google Drive share URL\n"
            "     (the long string between /d/ and /view) and add it as a GitHub secret."
        )

    # Create a credentials object using the service account info.
    # The "scope" tells Google what permissions this script needs.
    # "drive" scope = full read/write access to Drive files shared with this service account.
    creds = service_account.Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/drive"],
    )

    # Build and return the Drive API client
    return build("drive", "v3", credentials=creds)


def download_excel_from_drive(local_path: str):
    """
    Download the Excel file from Google Drive to the local GitHub Actions runner.

    WHY: The runner starts fresh every time with no files. We pull the latest
         Excel (which may have been manually updated by you in Drive) before
         processing so we always work with the most up-to-date ETF list.

    ARGS:
        local_path: The local filename to save the downloaded file as.
                    Uses CONFIG["excel_file"] = "etf_data_amfi1.xlsx"
    """
    log.info("Downloading Excel from Google Drive (file ID: %s) ...", DRIVE_FILE_ID)

    service = _get_drive_service()

    # Request the raw file bytes from Drive using the file ID
    request = service.files().get_media(fileId=DRIVE_FILE_ID)

    # Write the downloaded bytes directly to a local file
    with open(local_path, "wb") as f:
        downloader = MediaIoBaseDownload(f, request)
        done = False
        while not done:
            # Download in chunks; status.progress() gives 0.0 -> 1.0
            status, done = downloader.next_chunk()
            if status:
                log.info("  Download progress: %d%%", int(status.progress() * 100))

    log.info("Download complete -> %s", local_path)


def upload_excel_to_drive(local_path: str):
    """
    Upload the locally updated Excel file back to Google Drive, overwriting the original.

    WHY: After writing new data to the Excel, we push it back to Drive so you
         can open it from any device. The next run will download this updated
         version, preserving PE History and other accumulated data.

    ARGS:
        local_path: Path to the local Excel file to upload.
    """
    log.info("Uploading updated Excel to Google Drive (file ID: %s) ...", DRIVE_FILE_ID)

    service = _get_drive_service()

    # MediaFileUpload wraps the local file for the Drive API upload
    media = MediaFileUpload(
        local_path,
        # MIME type tells Drive this is an Excel file
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,  # resumable=True allows large file uploads in chunks
    )

    # files().update() overwrites the existing Drive file (same file ID, new content)
    service.files().update(
        fileId=DRIVE_FILE_ID,
        media_body=media,
    ).execute()

    log.info("Upload complete. Excel updated in Google Drive.")


# =============================================================================
# SECTION 2: EXCEL STYLING HELPERS
# =============================================================================
# These are reusable style objects and helper functions for making the Excel
# output look clean and professional.
# =============================================================================

# Header row style: dark blue background with white bold text
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # Dark navy blue
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)

# Data row style: plain Arial text
DATA_FONT = Font(name="Arial", size=10)

# Alternating row background: light blue for even rows, white for odd rows
# This "zebra striping" makes rows easier to read
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")  # Light blue

# Thin border around every cell for a clean grid look
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# Text alignment options
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def _style_header(cell, text):
    """
    Apply header styling to a cell and set its value.
    Used for the first row of every sheet.
    """
    cell.value     = text
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = BORDER_THIN


def _style_data(cell, value, alt_row=False):
    """
    Apply data row styling to a cell and set its value.
    alt_row=True applies the light-blue alternating background.
    """
    cell.value     = value
    cell.font      = DATA_FONT
    cell.fill      = ALT_FILL if alt_row else PatternFill()  # PatternFill() = no fill (white)
    cell.alignment = CENTER
    cell.border    = BORDER_THIN


def _autofit(ws, min_width=12, max_width=30):
    """
    Auto-resize all columns in a worksheet to fit their content.
    Caps width between min_width and max_width characters to prevent
    extremely narrow or extremely wide columns.
    """
    for col in ws.columns:
        # Find the longest cell value in this column
        length = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col
        )
        # Set column width with padding (+4) clamped to min/max
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            max(min_width, min(length + 4, max_width))


# =============================================================================
# SECTION 3: READING THE ETF LIST FROM EXCEL
# =============================================================================
# Reads the input ETF symbols and ISINs from the Excel file.
# This is what tells the script WHICH ETFs to monitor.
# =============================================================================

def load_etf_registry(excel_path: str) -> dict:
    """
    Read the list of ETFs from the Excel file and return them as a dictionary.

    RETURNS:
        A dict like:
        {
          "MAFANG": {
              "isin":          "INF769K01HF4",
              "name":          "Mirae Asset NYSE FANG+ ETF",
              "threshold_pct": 15.0,
              "row":           2        <- Excel row number, used when writing results back
          },
          "MON100": { ... },
          ...
        }

    If the Excel file doesn't exist, creates a template file with sample ETFs
    so the user knows the expected format.
    """
    p = Path(excel_path)

    # If the file doesn't exist locally (first run ever), create a template
    if not p.exists():
        _create_template_excel(excel_path)
        log.warning(
            "No Excel found locally — created template at %s. "
            "Populate it, upload to Drive, then re-run.", excel_path
        )
        return {}

    wb = openpyxl.load_workbook(excel_path)
    # Use the sheet named "ETF Data" if it exists, otherwise use the first sheet
    ws = wb[MAIN_SHEET] if MAIN_SHEET in wb.sheetnames else wb.active

    registry = {}
    # Start from row 2 (row 1 is the header)
    for row in range(2, ws.max_row + 1):
        symbol = ws.cell(row=row, column=COL_SYMBOL).value
        isin   = ws.cell(row=row, column=COL_ISIN).value

        # Skip rows where symbol or ISIN is missing
        if not symbol or not isin:
            continue

        symbol    = str(symbol).strip().upper()   # e.g. "mafang " -> "MAFANG"
        isin      = str(isin).strip()
        name      = ws.cell(row=row, column=COL_NAME).value or symbol  # fallback to symbol
        thr_v     = ws.cell(row=row, column=COL_THRESHOLD).value
        threshold = float(thr_v) if thr_v else CONFIG["default_threshold_pct"]

        registry[symbol] = {
            "isin":          isin,
            "name":          str(name).strip(),
            "threshold_pct": threshold,
            "row":           row,   # Save the row number so we can write results back later
        }

    log.info("Loaded %d ETFs from %s", len(registry), excel_path)
    return registry


def _create_template_excel(excel_path: str):
    """
    Create a blank Excel file with the correct structure and some sample ETF rows.
    This runs only on the very first time if no Excel exists.
    The user should fill in their own ETFs and re-upload to Drive.
    """
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = MAIN_SHEET

    # Write header row (fixed input columns + all output columns)
    headers = ["Symbol", "ISIN", "Name", "Alert Threshold (%)"] + OUTPUT_COLS
    for c, h in enumerate(headers, 1):
        _style_header(ws.cell(row=1, column=c), h)

    # Add 5 sample ETF rows so the user sees the expected format
    samples = [
        ("MAFANG",     "INF769K01HF4", "Mirae Asset NYSE FANG+ ETF",       15.0),
        ("MASPTOP50",  "INF769K01HP3", "Mirae Asset S&P 500 Top 50 ETF",   15.0),
        ("MAHKTECH",   "INF769K01HS7", "Mirae Asset Hang Seng TECH ETF",    15.0),
        ("HNGSNGBEES", "INF204KB19I1", "Nippon India ETF Hang Seng BeES",   15.0),
        ("MON100",     "INF247L01AP3", "Motilal Oswal NASDAQ 100 ETF",      15.0),
    ]
    for r, (sym, isin, name, thr) in enumerate(samples, 2):
        ws.cell(row=r, column=1).value = sym
        ws.cell(row=r, column=2).value = isin
        ws.cell(row=r, column=3).value = name
        ws.cell(row=r, column=4).value = thr

    _ensure_pe_sheets(wb)  # Also create the PE History and PE Comparison sheets
    _autofit(ws)
    wb.save(excel_path)


def _ensure_pe_sheets(wb: openpyxl.Workbook):
    """
    Add the PE History and PE Comparison sheets to the workbook if they don't exist.
    Called both on template creation and before writing results.

    PE History sheet   : Accumulates one row per ETF per run with its P/E ratio.
                         Builds up over time, enabling month-over-month comparison.

    PE Comparison sheet: Shows each ETF's current P/E vs its P/E from ~30 days ago,
                         with a colour-coded change column (green=cheaper, red=pricier).
    """
    if PE_HIST_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(PE_HIST_SHEET)
        for c, h in enumerate(["Run Date", "Symbol", "Name", "P/E Ratio"], 1):
            _style_header(ws.cell(row=1, column=c), h)

    if PE_CMP_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(PE_CMP_SHEET)
        for c, h in enumerate(
            ["Symbol", "Name", "P/E (This Run)", "P/E (Last Month)", "Change", "Change (%)"], 1
        ):
            _style_header(ws.cell(row=1, column=c), h)


# =============================================================================
# SECTION 4: WRITING RESULTS BACK TO EXCEL
# =============================================================================
# After fetching data for all ETFs, these functions write everything into
# the three Excel sheets: main data, PE history log, and PE comparison.
# =============================================================================

def write_results_to_excel(excel_path: str, registry: dict, results: dict):
    """
    Write all fetched ETF data into the local Excel file.

    ARGS:
        excel_path: Local path to the Excel file (already downloaded from Drive).
        registry:   The ETF list dict from load_etf_registry().
        results:    Dict of {symbol: {ltp, nav, diff_pct, pe, ...}} from fetch_all_data().
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[MAIN_SHEET] if MAIN_SHEET in wb.sheetnames else wb.active

    # Re-write all header cells (refreshes styling in case the file was manually edited)
    fixed_headers = ["Symbol", "ISIN", "Name", "Alert Threshold (%)"]
    all_headers   = fixed_headers + OUTPUT_COLS
    for c, h in enumerate(all_headers, 1):
        _style_header(ws.cell(row=1, column=c), h)
    ws.row_dimensions[1].height = 22   # Slightly taller header row
    ws.freeze_panes = "A2"             # Freeze header row so it stays visible when scrolling

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Timestamp for "Last Updated" column

    for symbol, meta in registry.items():
        row      = meta["row"]              # The Excel row number for this ETF
        r        = results.get(symbol, {})  # Fetched data for this ETF (empty dict if failed)
        alt      = (row % 2 == 0)           # True for even rows -> apply alternating background

        ltp      = r.get("ltp")
        high_52w = r.get("high_52w")
        low_52w  = r.get("low_52w")

        # Calculate % rise from 52W low  e.g. LTP=120, Low=100 -> rise = 20%
        rise_from_low = (
            round((ltp - low_52w) / low_52w * 100, 2)
            if ltp and low_52w else None
        )

        # Calculate % fall from 52W high  e.g. LTP=80, High=100 -> fall = 20%
        fall_from_high = (
            round((high_52w - ltp) / high_52w * 100, 2)
            if ltp and high_52w else None
        )

        # Build the list of values to write — ORDER MUST MATCH OUTPUT_COLS
        values = [
            _fmt(ltp,                    "Rs"),
            _fmt(r.get("expense_ratio"), "%"),
            _fmt(high_52w,               "Rs"),
            _fmt(low_52w,                "Rs"),
            _fmt(rise_from_low,          "%"),
            _fmt(fall_from_high,         "%"),
            _fmt(r.get("beta")),
            _fmt(r.get("peg")),
            _fmt(r.get("pe")),
            _fmt(r.get("volume"),        fmt="int"),
            _fmt(r.get("nav"),           "Rs"),
            r.get("nav_date", "N/A"),    # NAV date is already a string
            _fmt(r.get("diff_pct"),      "%"),
            now_str,
        ]

        # Style the fixed input columns (A-D) — don't overwrite their values
        for c in range(1, 5):
            cell           = ws.cell(row=row, column=c)
            cell.font      = DATA_FONT
            cell.fill      = ALT_FILL if alt else PatternFill()
            cell.alignment = LEFT
            cell.border    = BORDER_THIN

        # Write and style the output columns (E onwards)
        for i, val in enumerate(values):
            _style_data(ws.cell(row=row, column=COL_OUTPUT_START + i), val, alt)

    _autofit(ws)  # Auto-resize all columns to fit content

    # ── Write to the PE History sheet ─────────────────────────────────────
    # Appends one row per ETF with today's P/E ratio.
    # Over time this builds a historical log for trend analysis.
    _ensure_pe_sheets(wb)
    ws_hist  = wb[PE_HIST_SHEET]
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    hist_row = ws_hist.max_row + 1  # Start writing after the last existing row

    for symbol, r in results.items():
        pe = r.get("pe")
        if pe is None:
            continue  # Skip ETFs where P/E couldn't be fetched
        ws_hist.cell(row=hist_row, column=1).value = run_date
        ws_hist.cell(row=hist_row, column=2).value = symbol
        ws_hist.cell(row=hist_row, column=3).value = registry[symbol]["name"]
        ws_hist.cell(row=hist_row, column=4).value = pe
        hist_row += 1

    _autofit(ws_hist)

    # ── Write to the PE Comparison sheet ──────────────────────────────────
    _write_pe_comparison(wb, registry, results)

    wb.save(excel_path)
    log.info("Excel saved locally: %s", excel_path)


def _write_pe_comparison(wb, registry, results):
    """
    Refresh the PE Comparison sheet.

    For each ETF, compares:
      - Current P/E (this run)
      - P/E from the closest data point ~30 days ago (from PE History sheet)

    The "Change" cell is coloured:
      - Green  = P/E went DOWN  (ETF became cheaper relative to earnings — good)
      - Red    = P/E went UP    (ETF became more expensive — caution)
      - Black  = No change
    """
    ws_hist = wb[PE_HIST_SHEET]
    ws_cmp  = wb[PE_CMP_SHEET]

    # Clear old comparison data (but keep the header row 1)
    for row in ws_cmp.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Build a history dict: {symbol: [(run_datetime, pe_value), ...]}
    # This reads ALL historical data from the PE History sheet
    history: dict[str, list] = {}
    for row in ws_hist.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        run_dt, sym, _, pe_val = row[0], row[1], row[2], row[3]
        if sym and pe_val is not None:
            history.setdefault(sym, []).append((run_dt, float(pe_val)))

    today   = datetime.now()
    cmp_row = 2  # Start writing from row 2 (row 1 is the header)

    for symbol, meta in registry.items():
        current_pe    = results.get(symbol, {}).get("pe")
        # Find the PE entry closest to 30 days ago
        last_month_pe = _find_closest_pe(history.get(symbol, []), today, days_ago=30)

        change     = None
        change_pct = None
        if current_pe is not None and last_month_pe is not None:
            change     = round(current_pe - last_month_pe, 4)
            # Percentage change: (new - old) / old * 100
            change_pct = round(change / last_month_pe * 100, 2) if last_month_pe else None

        row_vals = [
            symbol,
            meta["name"],
            _fmt(current_pe),
            _fmt(last_month_pe),
            _fmt(change),
            _fmt(change_pct, "%"),
        ]
        alt = (cmp_row % 2 == 0)
        for c, v in enumerate(row_vals, 1):
            _style_data(ws_cmp.cell(row=cmp_row, column=c), v, alt)

        # Colour-code the Change cell (column 5)
        if change is not None:
            cell      = ws_cmp.cell(row=cmp_row, column=5)
            cell.font = Font(
                name="Arial", size=10, bold=True,
                color=(
                    "006100" if change < 0   # Dark green = P/E fell (cheaper)
                    else "9C0006" if change > 0  # Dark red = P/E rose (expensive)
                    else "000000"            # Black = no change
                ),
            )
        cmp_row += 1

    _autofit(ws_cmp)

    # Re-write headers (they were cleared when we blanked old data above)
    for c, h in enumerate(
        ["Symbol", "Name", "P/E (This Run)", "P/E (Last Month)", "Change", "Change (%)"], 1
    ):
        _style_header(ws_cmp.cell(row=1, column=c), h)


def _find_closest_pe(history: list, reference: datetime, days_ago: int):
    """
    Search through the PE history list and return the P/E value from the
    entry whose date is closest to (reference - days_ago).

    Example: reference = today (Apr 9), days_ago = 30
             target = Mar 10.  Find the PE reading closest to Mar 10.

    ARGS:
        history:   List of (datetime_string, pe_float) tuples for one symbol.
        reference: The reference datetime (usually today).
        days_ago:  How many days back to look.

    RETURNS:
        Float P/E value, or None if history is empty.
    """
    if not history:
        return None

    from datetime import timedelta
    # Target date = today minus 30 days (midnight)
    target = reference.replace(hour=0, minute=0, second=0) - timedelta(days=days_ago)

    best       = None
    best_delta = None

    for run_dt, pe in history:
        # Parse date strings like "2026-03-10 10:00" into datetime objects
        if isinstance(run_dt, str):
            try:
                dt = datetime.strptime(run_dt, "%Y-%m-%d %H:%M")
            except ValueError:
                continue   # Skip malformed dates
        else:
            dt = run_dt

        # Calculate how far this entry is from our target date (in seconds)
        delta = abs((dt - target).total_seconds())

        # Keep track of the entry with the smallest delta (closest to target)
        if best_delta is None or delta < best_delta:
            best_delta = delta
            best       = pe

    return best


def _fmt(val, unit="", fmt=""):
    """
    Format a value for display in Excel cells.

    ARGS:
        val:  The value to format (float, int, or None).
        unit: "Rs" for currency,  "%" for percentage,  "" for plain number.
        fmt:  "int" to format as a comma-separated integer (e.g. 1,234,567).

    RETURNS:
        A formatted string, or "N/A" if val is None.

    EXAMPLES:
        _fmt(163.45, "Rs")       -> "Rs163.4500"
        _fmt(15.23,  "%")        -> "15.23%"
        _fmt(-5.1,   "%")        -> "-5.10%"
        _fmt(1234567, fmt="int") -> "1,234,567"
        _fmt(None)               -> "N/A"
    """
    if val is None:
        return "N/A"
    if fmt == "int":
        return f"{int(val):,}"                          # e.g. 1234567 -> "1,234,567"
    if unit == "Rs":
        return f"Rs{val:,.4f}"                          # e.g. Rs163.4500
    if unit == "%":
        return f"{val:+.2f}%" if isinstance(val, float) and val < 0 else f"{val:.2f}%"
    return round(val, 4) if isinstance(val, float) else val


# =============================================================================
# SECTION 5: MARKET STATUS
# =============================================================================
# NSE (National Stock Exchange of India) is open Mon-Fri, 9:15 AM - 3:30 PM IST.
# GitHub Actions runs in UTC, so we convert to IST (UTC+5:30) before checking.
# Live iNAV is only available during market hours — outside hours we use
# the last declared NAV from AMFI instead.
# =============================================================================

# Set of NSE trading holidays (dates when market is closed despite being a weekday).
# Update this list each year with the official NSE holiday calendar.
_NSE_HOLIDAYS = {
    "2025-01-26", "2025-02-19", "2025-03-14", "2025-03-31",
    "2025-04-10", "2025-04-14", "2025-04-18", "2025-05-01",
    "2025-08-15", "2025-08-27", "2025-10-02", "2025-10-24",
    "2025-11-05", "2025-12-25",
    "2026-01-26", "2026-03-19", "2026-04-02", "2026-04-03",
    "2026-04-14", "2026-04-30", "2026-05-01", "2026-08-17",
    "2026-09-16", "2026-10-01", "2026-10-20", "2026-11-24",
    "2026-12-25",
}


def _ist_now() -> datetime:
    """
    Return the current datetime in IST (Indian Standard Time = UTC+5:30).
    Uses Python's zoneinfo module which is built into Python 3.9+.
    """
    from zoneinfo import ZoneInfo
    return datetime.now(ZoneInfo("Asia/Kolkata"))


def is_market_open() -> bool:
    """
    Return True if NSE is currently open for trading, False otherwise.

    Checks three conditions:
    1. Today is a weekday (Mon=0 ... Fri=4; Sat=5, Sun=6 are closed)
    2. Today is not an NSE trading holiday
    3. Current IST time is between 9:15 AM and 3:30 PM
    """
    now = _ist_now()
    if now.weekday() >= 5:                          # Saturday or Sunday
        return False
    if now.strftime("%Y-%m-%d") in _NSE_HOLIDAYS:  # Public/trading holiday
        return False
    t = now.time()
    return dtime(9, 15) <= t <= dtime(15, 30)       # Within trading hours


def market_status_label() -> str:
    """
    Return a human-readable string describing the current market status.
    Used in log messages and alert SMS/emails.
    """
    now = _ist_now()
    if now.weekday() >= 5:
        return "Weekend"
    if now.strftime("%Y-%m-%d") in _NSE_HOLIDAYS:
        return "Market Holiday"
    t = now.time()
    if t < dtime(9, 15):
        return "Pre-market"
    if t > dtime(15, 30):
        return "Post-market"
    return "Open"


# =============================================================================
# SECTION 6: NSE SESSION MANAGEMENT
# =============================================================================
# NSE's website uses anti-bot protection. Before calling any NSE API endpoint,
# we must first visit the NSE homepage to receive session cookies (nsit, nseappid).
# Without these cookies, the API returns 401 Unauthorized or empty JSON.
#
# We keep a single requests.Session alive and refresh its cookies every 30 minutes.
# =============================================================================

_NSE_SESSION     = None   # Stores the active requests.Session object
_NSE_SESSION_AT  = 0.0    # Unix timestamp of when the session was last refreshed
_NSE_SESSION_TTL = 1800   # 30 minutes in seconds — refresh after this time

_NSE_BASE    = "https://www.nseindia.com"

# Browser-like headers to avoid being blocked as a bot
_NSE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/123.0.0.0 Safari/537.36"
    ),
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "en-IN,en-US;q=0.9,en;q=0.8",
    "Referer":         "https://www.nseindia.com/",
}


def _get_nse_session() -> requests.Session:
    """
    Return a warmed-up NSE session with valid cookies.
    Refreshes the session if it is older than 30 minutes.

    Warm-up steps:
    1. GET NSE homepage -> sets the initial nsit / nseappid cookies
    2. GET the ETF market-data page -> fully primes the session
    Both requests include a small sleep to mimic human browsing behaviour.
    """
    global _NSE_SESSION, _NSE_SESSION_AT

    # Return existing session if it's still fresh (within TTL)
    if _NSE_SESSION and (time.time() - _NSE_SESSION_AT) < _NSE_SESSION_TTL:
        return _NSE_SESSION

    session = requests.Session()
    session.headers.update(_NSE_HEADERS)

    try:
        session.get(_NSE_BASE, timeout=15)               # Step 1: get initial cookies
        time.sleep(1)                                     # Wait 1 second (polite delay)
        session.get(                                      # Step 2: prime the session
            f"{_NSE_BASE}/market-data/exchange-traded-funds-etf", timeout=15
        )
        time.sleep(0.5)
        _NSE_SESSION    = session
        _NSE_SESSION_AT = time.time()
        log.debug("NSE session refreshed (cookies: %s)", list(session.cookies.keys()))
    except Exception as e:
        log.warning("NSE session warm-up failed: %s", e)
        _NSE_SESSION = session  # Use the session even if warm-up partially failed

    return _NSE_SESSION


# =============================================================================
# SECTION 7: LIVE iNAV FROM NSE
# =============================================================================
# iNAV = Indicative NAV — a real-time estimate of an ETF's fair value,
# updated every 15 seconds during market hours by the AMC's calculation agent.
#
# NSE provides this via their quote-equity API endpoint:
#   GET https://www.nseindia.com/api/quote-equity?symbol=MAFANG
#   Response -> data["metadata"]["iNavValue"]  (a string like "163.45")
#
# This is only populated during market hours. Outside hours, it is "0" or absent.
# =============================================================================

_NSE_INAV_SYMBOL_CACHE: dict = {}  # {symbol: (inav_value, fetched_at_timestamp)}
_NSE_INAV_SYMBOL_TTL = 60          # Cache for 60 seconds to avoid hammering NSE


def fetch_inav_nse(symbol: str):
    """
    Fetch the live iNAV for an ETF from NSE's quote-equity API.

    RETURNS:
        Float iNAV value if available and > 0, or None if:
        - Market is closed (iNavValue will be "0" or absent)
        - Symbol not found on NSE
        - Network error
    """
    # Return cached value if it was fetched within the last 60 seconds
    cached = _NSE_INAV_SYMBOL_CACHE.get(symbol.upper())
    if cached and (time.time() - cached[1]) < _NSE_INAV_SYMBOL_TTL:
        return cached[0] if cached[0] > 0 else None

    try:
        session  = _get_nse_session()
        url      = f"{_NSE_BASE}/api/quote-equity?symbol={symbol.upper()}"
        resp     = session.get(url, timeout=15)
        resp.raise_for_status()  # Raise an error for HTTP 4xx/5xx responses
        data     = resp.json()

        # The iNavValue can be in either of these two locations in the response JSON
        inav_raw = (
            data.get("metadata", {}).get("iNavValue")
            or data.get("priceInfo", {}).get("iNavValue")
        )

        if inav_raw is None:
            log.debug("NSE quote-equity: iNavValue absent for %s", symbol)
            return None

        # Clean up the string and check for zero/null values
        inav_str = str(inav_raw).replace(",", "").strip()
        if not inav_str or inav_str in ("-", "0", "0.0", "null", "NA", "N/A"):
            log.debug("NSE iNAV is zero/null for %s: %r", symbol, inav_raw)
            return None

        inav = float(inav_str)
        if inav <= 0:
            return None

        # Cache the result
        _NSE_INAV_SYMBOL_CACHE[symbol.upper()] = (inav, time.time())
        log.debug("NSE iNAV  symbol=%-12s  iNav=%.4f", symbol, inav)
        return inav

    except Exception as e:
        log.warning("NSE iNAV fetch failed for %s: %s", symbol, e)
        # Force session refresh on the next call (cookies may have expired)
        global _NSE_SESSION_AT
        _NSE_SESSION_AT = 0.0
        return None


# =============================================================================
# SECTION 8: EXPENSE RATIO FETCHING
# =============================================================================
# Expense ratio (TER — Total Expense Ratio) is the annual cost of holding an ETF,
# expressed as a percentage of AUM (e.g. 0.59 means 0.59% per year).
#
# Two free data sources are tried in order:
#
# PRIMARY:  mfdata.in/api/v1/schemes/{amfi_code}
#   - Modern, well-maintained API with daily updates from AMFI
#   - Requires first resolving ISIN -> AMFI scheme code via NAVAll.txt
#
# FALLBACK: mf.captnemo.in/search?q={isin}
#   - Older API, works for most ETFs but occasionally misses some
#
# TER changes at most once or twice a year, so we cache it for 24 hours.
# =============================================================================

_EXPENSE_CACHE:   dict = {}   # {isin: expense_ratio_float_or_None}
_EXPENSE_FETCHED: dict = {}   # {isin: timestamp}
_EXPENSE_TTL = 86400           # 24 hours in seconds

_MFDATA_BASE  = "https://mfdata.in/api/v1/schemes"
_CAPTNEMO_URL = "https://mf.captnemo.in/search"

# Generic browser-like headers for requests that don't need NSE session cookies
_GENERIC_HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/123.0.0.0",
    "Accept-Language": "en-IN,en-US;q=0.9",
}


def fetch_expense_ratio(isin: str):
    """
    Return the expense ratio (%) for an ETF identified by its ISIN.
    Tries mfdata.in first, then mf.captnemo.in as fallback.
    Returns None if both sources fail.
    """
    # Return cached value if within 24 hours (TER rarely changes)
    if isin in _EXPENSE_CACHE:
        if (time.time() - _EXPENSE_FETCHED.get(isin, 0)) < _EXPENSE_TTL:
            return _EXPENSE_CACHE[isin]

    val = _fetch_expense_mfdata(isin)        # Try primary source
    if val is None:
        val = _fetch_expense_captnemo(isin)  # Try fallback source

    _EXPENSE_CACHE[isin]   = val
    _EXPENSE_FETCHED[isin] = time.time()
    return val


def _fetch_expense_mfdata(isin: str):
    """
    Fetch expense ratio from mfdata.in using the AMFI scheme code.
    Steps:
    1. Resolve ISIN -> AMFI scheme code (via NAVAll.txt)
    2. Call GET /api/v1/schemes/{scheme_code}
    3. Read the "expense_ratio" field from the response
    """
    scheme_code = _scheme_code_for_isin(isin)
    if not scheme_code:
        return None
    try:
        r    = requests.get(
            f"{_MFDATA_BASE}/{scheme_code}", headers=_GENERIC_HEADERS, timeout=15
        )
        r.raise_for_status()
        body = r.json()
        if body.get("status") == "success":
            raw = body.get("data", {}).get("expense_ratio")
            if raw not in (None, "", "null"):
                val = float(raw)
                if val > 0:
                    return val
    except Exception as e:
        log.debug("mfdata.in expense fetch failed (ISIN=%s): %s", isin, e)
    return None


def _fetch_expense_captnemo(isin: str):
    """
    Fetch expense ratio from mf.captnemo.in by searching with the ISIN.
    The API returns a JSON array of matching schemes; we find the one
    whose ISIN matches exactly and read its expense_ratio field.
    """
    try:
        r     = requests.get(
            _CAPTNEMO_URL, params={"q": isin}, headers=_GENERIC_HEADERS, timeout=15
        )
        r.raise_for_status()
        items = r.json()
        if isinstance(items, list):
            for item in items:
                # captnemo uses uppercase "ISIN" key — check both cases to be safe
                if item.get("ISIN") == isin or item.get("isin") == isin:
                    raw = item.get("expense_ratio") or item.get("expenseRatio")
                    if raw not in (None, "", "null", "0", 0):
                        return float(raw)
    except Exception as e:
        log.debug("captnemo expense fetch failed (ISIN=%s): %s", isin, e)
    return None


# =============================================================================
# SECTION 9: AMFI NAVAll.txt — DECLARED NAV AND SCHEME CODE RESOLUTION
# =============================================================================
# AMFI (Association of Mutual Funds in India) publishes a daily text file
# with the latest declared NAV for every mutual fund and ETF.
#
# URL: https://www.amfiindia.com/spages/NAVAll.txt
#
# File format (semicolon-separated):
#   SchemeCode ; ISIN_Growth ; ISIN_DivReinvest ; SchemeName ; NAV ; Date
#
# We use this file for two purposes:
#   1. Getting the declared NAV when live iNAV is unavailable (outside market hours)
#   2. Resolving an ISIN to an AMFI scheme code (needed for mfapi and mfdata calls)
#
# The file is cached for 10 minutes to avoid downloading it repeatedly.
# =============================================================================

_AMFI_URL        = "https://www.amfiindia.com/spages/NAVAll.txt"
_AMFI_TEXT       = None    # Cached raw text of the file
_AMFI_FETCHED_AT = 0.0     # When it was last fetched
_AMFI_TTL        = 600     # 10 minutes in seconds

# Resolved ISIN -> scheme code mappings (cached for the life of the script run)
_ISIN_TO_SCHEME_CODE: dict = {}


def _get_amfi_text():
    """
    Fetch and cache the AMFI NAVAll.txt file.
    Returns the raw text, or None if the download fails.
    """
    global _AMFI_TEXT, _AMFI_FETCHED_AT
    if _AMFI_TEXT and (time.time() - _AMFI_FETCHED_AT) < _AMFI_TTL:
        return _AMFI_TEXT   # Return cached version
    try:
        r = requests.get(_AMFI_URL, headers=_GENERIC_HEADERS, timeout=25)
        r.raise_for_status()
        _AMFI_TEXT       = r.text
        _AMFI_FETCHED_AT = time.time()
        log.debug("AMFI NAVAll.txt refreshed (%d bytes)", len(_AMFI_TEXT))
    except Exception as e:
        log.warning("AMFI fetch failed: %s", e)
    return _AMFI_TEXT


def _parse_amfi_line(isin: str):
    """
    Search through NAVAll.txt for a line matching the given ISIN.

    Each line looks like:
      120503;INF769K01HF4;INF769K01HG2;Mirae Asset NYSE FANG+ ETF;163.4500;08-Apr-2026

    ISIN can be in position [1] (growth) or [2] (dividend reinvestment).

    RETURNS:
        Tuple (scheme_code, nav, date) if found, or None if not found.
    """
    text = _get_amfi_text()
    if not text:
        return None
    for line in text.splitlines():
        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 6:
            continue   # Skip malformed lines (headers, blank lines, etc.)
        # Check if either ISIN column matches
        if parts[1] == isin or parts[2] == isin:
            try:
                nav = float(parts[4].replace(",", ""))
                if nav > 0:
                    return parts[0], nav, parts[5]   # (scheme_code, nav, date)
            except (ValueError, IndexError):
                pass
    return None


def fetch_nav_amfi_by_isin(isin: str):
    """
    Get the latest declared NAV for an ETF from AMFI NAVAll.txt.
    Used as the final fallback when NSE iNAV and mfapi both fail.

    RETURNS:
        Tuple (nav_float, date_string) or None if ISIN not found.
    """
    result = _parse_amfi_line(isin)
    if result:
        _, nav, date = result
        return nav, date
    log.warning("ISIN '%s' not found in AMFI NAVAll.txt", isin)
    return None


def _scheme_code_for_isin(isin: str):
    """
    Resolve an ISIN to an AMFI scheme code using NAVAll.txt.
    The scheme code is a 6-digit number used by mfapi.in and mfdata.in.

    Results are cached for the life of the script run to avoid repeated parsing.

    RETURNS:
        Scheme code string (e.g. "120503") or None if not found.
    """
    if isin in _ISIN_TO_SCHEME_CODE:
        return _ISIN_TO_SCHEME_CODE[isin]   # Return cached result
    result = _parse_amfi_line(isin)
    if result:
        code = result[0]
        _ISIN_TO_SCHEME_CODE[isin] = code   # Cache for this run
        return code
    return None


# =============================================================================
# SECTION 10: mfapi.in — LATEST DECLARED NAV
# =============================================================================
# mfapi.in is a free, community-maintained API for Indian mutual fund NAVs.
# It provides the latest declared NAV by AMFI scheme code.
#
# Used as the first declared-NAV fallback (before raw AMFI text parsing).
# Cached for 10 minutes since NAV is declared once a day after market close.
# =============================================================================

_MFAPI_BASE  = "https://api.mfapi.in/mf"
_MFAPI_CACHE: dict = {}   # {scheme_code: {nav, date, fetched_at}}
_MFAPI_TTL   = 600        # 10 minutes


def fetch_nav_mfapi(isin: str):
    """
    Fetch the latest declared NAV from mfapi.in using the AMFI scheme code.

    Flow:
    1. Resolve ISIN -> scheme code (via _scheme_code_for_isin)
    2. Call GET https://api.mfapi.in/mf/{scheme_code}/latest
    3. Parse the NAV and date from the response

    RETURNS:
        Tuple (nav_float, date_string) or None on failure.
    """
    scheme_code = _scheme_code_for_isin(isin)
    if not scheme_code:
        return None   # ISIN not found in AMFI data

    # Return cached result if fresh
    cached = _MFAPI_CACHE.get(scheme_code)
    if cached and (time.time() - cached["fetched_at"]) < _MFAPI_TTL:
        return cached["nav"], cached["date"]

    try:
        r    = requests.get(
            f"{_MFAPI_BASE}/{scheme_code}/latest",
            headers=_GENERIC_HEADERS, timeout=15,
        )
        r.raise_for_status()
        body = r.json()
        data = body.get("data", [])
        if data:
            nav  = float(data[0]["nav"])
            date = data[0]["date"]
            _MFAPI_CACHE[scheme_code] = {"nav": nav, "date": date, "fetched_at": time.time()}
            log.debug("mfapi NAV  scheme=%s  nav=%.4f  date=%s", scheme_code, nav, date)
            return nav, date
    except Exception as e:
        log.warning("mfapi fetch failed for scheme %s: %s", scheme_code, e)
    return None


# =============================================================================
# SECTION 11: BSE LIVE iNAV  (secondary live fallback)
# =============================================================================
# BSE (Bombay Stock Exchange) also publishes live iNAV for ETFs during
# market hours via their API. This is used as a backup if NSE's iNAV
# endpoint returns nothing.
#
# The BSE API returns a list of ALL ETFs' iNAVs in one call, so we fetch
# once and scan the list for the matching symbol. Cached for 60 seconds.
# =============================================================================

_BSE_INAV_CACHE:      list  = []   # Full list of ETF iNAVs from BSE
_BSE_INAV_FETCHED_AT: float = 0.0  # When the list was last fetched
_BSE_INAV_TTL = 60                  # Cache for 60 seconds


def _get_bse_inav_list() -> list:
    """
    Fetch and cache the full BSE iNAV list.
    Returns an empty list if the request fails.
    """
    global _BSE_INAV_CACHE, _BSE_INAV_FETCHED_AT
    if _BSE_INAV_CACHE and (time.time() - _BSE_INAV_FETCHED_AT) < _BSE_INAV_TTL:
        return _BSE_INAV_CACHE
    try:
        r = requests.get(
            "https://api.bseindia.com/BseIndiaAPI/api/ETFiNav/w",
            headers={"User-Agent": "Mozilla/5.0", "Referer": "https://www.bseindia.com/"},
            timeout=15,
        )
        if r.status_code == 200:
            _BSE_INAV_CACHE      = r.json()
            _BSE_INAV_FETCHED_AT = time.time()
    except Exception as e:
        log.warning("BSE iNAV fetch failed: %s", e)
    return _BSE_INAV_CACHE


def _fetch_inav_bse_for_symbol(symbol: str):
    """
    Search the BSE iNAV list for an ETF matching the given NSE symbol.
    Matching is done by checking if the NSE symbol appears in the BSE scheme name.

    RETURNS:
        Float iNAV value if found and > 0, or None.
    """
    for item in _get_bse_inav_list():
        name = str(item.get("scname", "")).upper()
        if symbol.upper() in name:
            # Try different field name variations (BSE API is not fully consistent)
            val = item.get("inav") or item.get("iNav") or item.get("nav")
            if val:
                try:
                    fval = float(str(val).replace(",", ""))
                    if fval > 0:
                        return fval
                except ValueError:
                    pass
    return None


# =============================================================================
# SECTION 12: YAHOO FINANCE — LTP, 52W HIGH/LOW, BETA, P/E, VOLUME
# =============================================================================
# yfinance is a Python library that wraps Yahoo Finance's unofficial API.
# NSE symbols are accessed by appending ".NS" (e.g. "MAFANG.NS").
#
# fast_info  : A lightweight object with key metrics (faster than ticker.info)
# ticker.info: Full metadata dict (slower, more fields)
# history()  : Historical OHLCV data (used as LTP fallback)
# =============================================================================

def fetch_yahoo_metrics(symbol: str) -> dict:
    """
    Fetch market metrics for an NSE ETF from Yahoo Finance.

    RETURNS a dict with keys:
        ltp      - Last Traded Price (most recent close or live price)
        high_52w - 52-week high price
        low_52w  - 52-week low price
        beta     - Beta coefficient (market volatility measure)
        peg      - PEG ratio (always None for Indian ETFs — not available)
        pe       - P/E ratio (trailing or forward)
        volume   - Today's traded volume (units)
    """
    ticker = yf.Ticker(f"{symbol}.NS")  # ".NS" suffix tells Yahoo this is an NSE symbol
    info   = ticker.info or {}          # Full info dict (can be empty on network issues)
    fi     = ticker.fast_info           # Fast info object

    # LTP: Try fast_info first (quicker), fall back to last row of price history
    ltp = getattr(fi, "last_price", None)
    if not ltp:
        hist = ticker.history(period="5d", interval="1d")
        ltp  = float(hist["Close"].iloc[-1]) if not hist.empty else None

    # 52-week high and low
    high_52w = getattr(fi, "year_high", None) or info.get("fiftyTwoWeekHigh")
    low_52w  = getattr(fi, "year_low",  None) or info.get("fiftyTwoWeekLow")

    # Volume: try multiple field names in priority order
    volume = (
        getattr(fi, "last_volume", None)
        or info.get("regularMarketVolume")
        or getattr(fi, "three_month_average_volume", None)
        or info.get("averageVolume")
    )

    # P/E ratio: try trailing first, then forward
    pe   = info.get("trailingPE") or info.get("forwardPE")

    # Beta: how much the ETF moves relative to the broader market
    beta = info.get("beta") or info.get("beta3Year")

    return {
        "ltp":      float(ltp)      if ltp      else None,
        "high_52w": float(high_52w) if high_52w else None,
        "low_52w":  float(low_52w)  if low_52w  else None,
        "beta":     float(beta)     if beta     else None,
        "peg":      None,   # PEG not available for Indian ETFs from any free source
        "pe":       float(pe)       if pe       else None,
        "volume":   int(volume)     if volume   else None,
    }


# =============================================================================
# SECTION 13: COMBINED DATA FETCH FOR ONE ETF
# =============================================================================
# This is the main data-gathering function that orchestrates all the sources
# above to build a complete data snapshot for one ETF.
#
# iNAV / NAV SOURCE PRIORITY (most live -> least live):
#   During market hours:
#     1. NSE quote-equity API  (live iNAV, updated every ~15 seconds)
#     2. BSE iNAV API          (live iNAV, secondary)
#   Any time:
#     3. mfapi.in              (declared NAV, updated once daily after market close)
#     4. AMFI NAVAll.txt       (declared NAV, raw source, final fallback)
# =============================================================================

def fetch_all_data(symbol: str, isin: str) -> dict:
    """
    Fetch all available data for one ETF and return it as a single dict.

    ARGS:
        symbol: NSE symbol (e.g. "MAFANG")
        isin:   ISIN code  (e.g. "INF769K01HF4")

    RETURNS:
        Dict with all metrics:
        {ltp, high_52w, low_52w, beta, peg, pe, volume,
         expense_ratio, nav, nav_date, diff_pct, is_live}
    """
    # Fetch price/volume/valuation metrics from Yahoo Finance
    metrics       = fetch_yahoo_metrics(symbol)
    # Fetch annual expense ratio
    expense_ratio = fetch_expense_ratio(isin)
    ltp           = metrics.get("ltp")

    nav, nav_date, is_live = None, "N/A", False

    # ── Live iNAV sources (only checked during market hours) ──────────────
    if is_market_open():
        nav = fetch_inav_nse(symbol)        # Source 1: NSE API
        if nav:
            nav_date, is_live = "live (NSE iNAV)", True

        if not nav:
            nav = _fetch_inav_bse_for_symbol(symbol)  # Source 2: BSE API
            if nav:
                nav_date, is_live = "live (BSE iNAV)", True

    # ── Declared NAV sources (used outside hours or if live iNAV failed) ──
    if not nav:
        result = fetch_nav_mfapi(isin)      # Source 3: mfapi.in
        if result:
            nav, nav_date = result
            nav_date = f"declared {nav_date} (mfapi)"

    if not nav:
        result = fetch_nav_amfi_by_isin(isin)  # Source 4: AMFI NAVAll.txt
        if result:
            nav, nav_date = result
            nav_date = f"declared {nav_date} (AMFI)"

    # ── Calculate LTP vs NAV difference ──────────────────────────────────
    # Positive diff = LTP trading at a PREMIUM to NAV (you're overpaying)
    # Negative diff = LTP trading at a DISCOUNT to NAV (potential opportunity)
    diff_pct = round((ltp - nav) / nav * 100, 4) if ltp and nav else None

    return {
        **metrics,           # Unpack all Yahoo metrics into this dict
        "expense_ratio": expense_ratio,
        "nav":           nav,
        "nav_date":      nav_date,
        "diff_pct":      diff_pct,
        "is_live":       is_live,
    }


# =============================================================================
# SECTION 14: ALERTS (EMAIL + SMS)
# =============================================================================
# An alert fires when abs(diff_pct) <= threshold_pct for an ETF.
# For example, if threshold = 15% and diff = -12% (ETF at 12% discount),
# an alert is sent because the discount is within the threshold window.
#
# Both Email (via Gmail SMTP) and SMS (via Twilio) are sent together.
# =============================================================================

def _alert_body(symbol, meta, r):
    """
    Build the plain-text alert message body.
    Used for both the email body and a shortened version for SMS.
    """
    ltp, nav  = r.get("ltp"), r.get("nav")
    diff_pct  = r.get("diff_pct", 0)
    direction = "DISCOUNT" if diff_pct and diff_pct < 0 else "PREMIUM"
    return (
        f"ETF Alert: {symbol}  -  {meta['name']}\n"
        f"  NAV/iNAV  : Rs{nav:.4f}  ({r.get('nav_date','N/A')})\n"
        f"  LTP       : Rs{ltp:.4f}\n"
        f"  Diff      : {diff_pct:+.2f}% ({direction})\n"
        f"  P/E       : {r.get('pe','N/A')}\n"
        f"  52W H/L   : Rs{r.get('high_52w','N/A')} / Rs{r.get('low_52w','N/A')}\n"
        f"  Market    : {market_status_label()}\n"
        f"  Time      : {datetime.now().strftime('%Y-%m-%d %H:%M:%S IST')}\n"
    )


def send_email(subject, body):
    """
    Send an alert email via Gmail using SMTP with STARTTLS encryption.

    Requires a Gmail App Password (not your regular Gmail password).
    To create one: Google Account -> Security -> 2-Step Verification -> App Passwords
    """
    cfg = CONFIG["email"]
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    msg = MIMEMultipart()
    msg["From"]    = cfg["sender_email"]
    msg["To"]      = cfg["recipient_email"]
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    # Connect to Gmail's SMTP server and send
    with smtplib.SMTP(cfg["smtp_host"], cfg["smtp_port"]) as srv:
        srv.ehlo()       # Identify ourselves to the server
        srv.starttls()   # Upgrade connection to encrypted TLS
        srv.ehlo()       # Re-identify after encryption
        srv.login(cfg["sender_email"], cfg["sender_password"])
        srv.sendmail(cfg["sender_email"], cfg["recipient_email"], msg.as_string())
    log.info("Email sent")


def send_sms(body):
    """
    Send an alert SMS via Twilio's REST API.

    Twilio provides a free trial account with a sandbox phone number.
    Sign up at twilio.com and get your Account SID, Auth Token, and a Twilio number.
    """
    cfg = CONFIG["sms"]
    r = requests.post(
        # Twilio's Messages endpoint for this account
        f"https://api.twilio.com/2010-04-01/Accounts/{cfg['account_sid']}/Messages.json",
        # HTTP Basic Auth: account_sid as username, auth_token as password
        auth=(cfg["account_sid"], cfg["auth_token"]),
        data={
            "From": cfg["from_number"],   # Your Twilio number
            "To":   cfg["to_number"],     # Your mobile number
            "Body": body,
        },
        timeout=10,
    )
    r.raise_for_status()
    log.info("SMS sent (SID: %s)", r.json().get("sid"))


def send_alert(symbol, meta, r):
    """
    Send both email and SMS alerts for an ETF that has met the threshold.
    Each alert channel is tried independently — if email fails, SMS still sends.
    """
    diff_pct  = r.get("diff_pct", 0) or 0
    direction = "discount" if diff_pct < 0 else "premium"
    body      = _alert_body(symbol, meta, r)
    subject   = f"[ETF Alert] {symbol} {abs(diff_pct):.2f}% {direction} to NAV"

    if CONFIG["email"]["enabled"]:
        try:
            send_email(subject, body)
        except Exception as e:
            log.error("Email failed: %s", e)

    if CONFIG["sms"]["enabled"]:
        try:
            nav, ltp = r.get("nav", 0), r.get("ltp", 0)
            # SMS has a 160-character limit so keep the message short
            send_sms(
                f"ETF {symbol}: NAV={nav:.2f} LTP={ltp:.2f} "
                f"Diff={diff_pct:+.2f}% ({direction})"
            )
        except Exception as e:
            log.error("SMS failed: %s", e)


# =============================================================================
# SECTION 15: LOCAL CACHE
# =============================================================================
# After fetching data for each ETF, we save a snapshot to a local JSON file
# (etf_cache.json) on the GitHub Actions runner.
#
# WHY: Useful for debugging — if the script crashes, you can inspect what
# data was successfully fetched before the failure. The file does NOT persist
# between GitHub Actions runs (the runner is fresh each time).
# =============================================================================

def save_cache(symbol, data):
    """
    Save/update the cached data for one ETF symbol in the local JSON cache file.
    Adds a "timestamp" field to record when this data was fetched.
    """
    p     = Path(CONFIG["cache_file"])
    cache = {}

    # Load existing cache if it exists
    if p.exists():
        try:
            cache = json.loads(p.read_text())
        except Exception:
            pass   # If the file is corrupt, start fresh

    # Add/update this symbol's entry with a timestamp
    cache[symbol.upper()] = {**data, "timestamp": datetime.now().isoformat()}

    try:
        p.write_text(json.dumps(cache, indent=2))
    except Exception as e:
        log.warning("Cache write failed: %s", e)


# =============================================================================
# SECTION 16: MAIN FUNCTION
# =============================================================================
# This is the entry point — the function that orchestrates all steps in order.
# GitHub Actions calls this once per scheduled run, then the script exits.
#
# EXECUTION FLOW:
#   1. Download Excel from Google Drive
#   2. Load ETF list from Excel
#   3. Fetch live data for every ETF
#   4. Send alerts for ETFs that meet the threshold
#   5. Write all results into the Excel file
#   6. Upload updated Excel back to Google Drive
# =============================================================================

def main():
    excel_path = CONFIG["excel_file"]   # Local filename (e.g. "etf_data_amfi1.xlsx")

    # Print a clear header in the GitHub Actions log
    log.info("=" * 60)
    log.info("Multi-ETF Monitor — GitHub Actions single-cycle run")
    log.info("Market status : %s", market_status_label())
    log.info("Run time (UTC): %s", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"))
    log.info("=" * 60)

    # ── STEP 1: Download the Excel from Google Drive ──────────────────────
    # The GitHub Actions runner starts fresh with no files, so we must
    # download the latest Excel first before we can read or write it.
    try:
        download_excel_from_drive(excel_path)
    except RuntimeError as e:
        # RuntimeError here means a configuration problem (missing secret, wrong ID, etc.)
        # Print the full error and stop — no point continuing without the input file.
        log.error("FATAL - could not download Excel from Drive:\n%s", e)
        raise SystemExit(1)   # Exit with code 1 -> GitHub Actions marks the run as FAILED

    # ── STEP 2: Load the ETF list from the downloaded Excel ───────────────
    registry = load_etf_registry(excel_path)
    if not registry:
        log.error("No ETFs found in Excel — nothing to process. Aborting.")
        raise SystemExit(1)

    # Track consecutive errors per symbol for logging purposes
    consecutive_errors: dict[str, int] = {s: 0 for s in registry}

    # ── STEP 3: Fetch data for every ETF ─────────────────────────────────
    results: dict[str, dict] = {}   # Will hold {symbol: fetched_data_dict}

    for symbol, meta in registry.items():
        try:
            # Fetch all metrics for this ETF (Yahoo + NSE/BSE/AMFI + expense ratio)
            r = fetch_all_data(symbol, meta["isin"])
            save_cache(symbol, r)          # Save to local JSON cache for debugging
            results[symbol] = r
            consecutive_errors[symbol] = 0

            # Log a one-line summary for this ETF
            ltp, nav, diff_pct = r.get("ltp"), r.get("nav"), r.get("diff_pct")
            log.info(
                "  %-14s  LTP=%-12s  NAV=%-12s  Diff=%s",
                symbol,
                f"Rs{ltp:.4f}"      if ltp      is not None else "N/A",
                f"Rs{nav:.4f}"      if nav      is not None else "N/A",
                f"{diff_pct:+.2f}%" if diff_pct is not None else "N/A",
            )

            # ── ALERT CHECK ───────────────────────────────────────────────
            # An alert fires when the LTP/NAV difference is WITHIN the threshold.
            # Example: threshold=15%, diff=-12% -> abs(-12) = 12 <= 15 -> ALERT
            # This means the ETF is close enough to NAV to act on.
            threshold = meta["threshold_pct"]
            if diff_pct is not None and abs(diff_pct) <= threshold:
                log.info(
                    "  %-14s  Threshold met (%.2f%% <= %.2f%%) — alerting!",
                    symbol, abs(diff_pct), threshold
                )
                send_alert(symbol, meta, r)

        except Exception as e:
            # Log the error but continue processing other ETFs
            consecutive_errors[symbol] += 1
            log.error("  %-14s  ERROR: %s", symbol, e)
            results[symbol] = {}   # Empty dict so write step skips this symbol gracefully

    # ── STEP 4: Write all results into the local Excel file ───────────────
    # This updates the "ETF Data", "PE History", and "PE Comparison" sheets.
    try:
        write_results_to_excel(excel_path, registry, results)
    except Exception as e:
        log.error("FATAL - Excel write failed: %s", e)
        raise SystemExit(1)

    # ── STEP 5: Upload the updated Excel back to Google Drive ─────────────
    # Overwrites the same file in Drive so you always have the latest data.
    # The next run will download THIS updated file, preserving PE history etc.
    try:
        upload_excel_to_drive(excel_path)
    except Exception as e:
        log.error("FATAL - could not upload Excel to Drive: %s", e)
        raise SystemExit(1)

    # ── Done ──────────────────────────────────────────────────────────────
    log.info("=" * 60)
    log.info("Cycle complete. Excel updated in Google Drive.")
    log.info("Next run scheduled by GitHub Actions cron.")
    log.info("=" * 60)


# ─────────────────────────────────────────────────────────────────────────────
# SCRIPT ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
# This block runs only when the script is executed directly:
#   python multi_etf_data_gdrive.py
#
# It does NOT run when the file is imported as a module by another script.
# This is a Python convention — always include this guard at the bottom.
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    main()
